from openpyxl import load_workbook
from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404

from student.models import student

def write(file):
	wb = load_workbook(filename = file)
	sheet = wb.sheetnames
	ws = wb[sheet[0]]
	# print(wb)
	# print(ws)
	# print(wb.get_sheet_names())
	# for row in ws.iter_rows():
	# 	for cell in row:
	# 		print(cell.value)
	username=[]
	password=[]
	firstname=[]
	lastname=[]
	middle_name=[]
	email=[]
	batch=[]
	branch=[]
	address=[]
	contact=[]
	roll_no=[]
	print('/n')
	for col in ws.iter_cols():
		for cell in col:
			# print(cell.value)
			if str(cell.value).strip() == 'username':
				for cell1 in col:
					if str(cell1.value).strip() == 'username':
						print(cell1.value)
						pass
					else:
						print(cell1.value)
						username.append(cell1.value)
			if str(cell.value).strip() == 'Password':
				for cell1 in col:
					if str(cell1.value).strip() == 'Password':
						pass
					else:
						password.append(cell1.value)
			if str(cell.value).strip() == 'first_name':
				for cell1 in col:
					if str(cell1.value).strip() == 'first_name':
						pass
					else:
						firstname.append(cell1.value)
			if str(cell.value).strip() == 'last_name':
				for cell1 in col:
					if str(cell1.value).strip() == 'last_name':
						pass
					else:
						lastname.append(cell1.value)
			if str(cell.value).strip() == 'email':
				for cell1 in col:
					if str(cell1.value).strip() == 'email':
						pass
					else:
						email.append(cell1.value)
			if str(cell.value).strip() == 'middle_name':
				for cell1 in col:
					if str(cell1.value).strip() == 'middle_name':
						pass
					else:
						middle_name.append(cell1.value)

			if str(cell.value).strip() == 'batch':
				for cell1 in col:
					if str(cell1.value).strip() == 'batch':
						pass
					else:
						batch.append(cell1.value)

			if str(cell.value).strip() == 'contact':
				for cell1 in col:
					if str(cell1.value).strip() == 'contact':
						pass
					else:
						contact.append(cell1.value)

			if str(cell.value).strip() == 'address':
				for cell1 in col:
					if str(cell1.value).strip() == 'address':
						pass
					else:
						address.append(cell1.value)

			if str(cell.value).strip() == 'branch':
				for cell1 in col:
					if str(cell1.value).strip() == 'branch':
						pass
					else:
						branch.append(cell1.value)

			if str(cell.value).strip() == 'roll_no':
				for cell1 in col:
					if str(cell1.value).strip() == 'roll_no':
						pass
					else:
						roll_no.append(cell1.value)

	print(username,roll_no,email,password,firstname,lastname,middle_name,batch,branch,contact,address)
	from django.contrib.auth.models import Group
	group = Group.objects.get(name='student')
	for i in range(0, len(username)):
		try:
			print(username[i])
			user = User.objects.get(username = username[i].strip())
		except User.DoesNotExist:
			user = User.objects.create_user(username=username[i], email=email[i], password=password[i],first_name=firstname[i],last_name=lastname[i] )
			group.user_set.add(user.id)
		except ValueError:
			pass
		try:
			user = User.objects.get(username=username[i])
			stu = student.objects.get(username=user)
		except:
			user = User.objects.get(username=username[i])
			stu = student.objects.get_or_create(username=user,roll_no=roll_no[i], email=email[i], first_name=firstname[i],last_name=lastname[i], middle_name=middle_name[i], batch=batch[i], branch=branch[i], contact=contact[i], address = address[i])
		